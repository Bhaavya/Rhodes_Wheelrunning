import openpyxl
import xlrd
import sys
import os
import datetime
import math
def main(argv,ofile,interval):
 f=-1

 if (interval.lower()).find('minute')!=-1:
  t=1
 
  index=(interval.lower()).find('minute')
 elif (interval.lower()).find('hour')!=-1:
  t=60
  
  index=(interval.lower()).find('hour')

 elif (interval.lower()).find('day')!=-1:
   t=1440
   index=(interval.lower()).find('day')
 
 interval=interval[:index]
 interval=interval.strip()
 interval=(int)(interval)
 id=list()
 animal=list()
 from_time=list()
 notes=list()
 notes1=list()
 notes2=list()
 to_time=list()
 flag=list()
 flag1=list()
 global_datetime=0
 global_time=0
 global_datetime=0
 max=0
 pc=0
 start_date=list()
 start_time=list()
 start_datetime=list()
 end_date=list()
 end_time=list()
 end_datetime=list()
 diff=list()
 sum=list()
 for ifile in argv:
    f=f+1
    
    last_start=list()
    last_end=list()
    wb=xlrd.open_workbook(ifile)
    ws=wb.sheet_by_index(0)
    
    i=0
    
        
    for row in range(0,ws.nrows):
      if ws.cell_value(row,1)=='':
          i=i+1
      else:
          break
    pc=ws.ncols/3
    if f==0:
       global_date=xlrd.xldate_as_tuple(ws.cell_value(i+1,0),wb.datemode)
       global_time=xlrd.xldate_as_tuple(ws.cell_value(i+1,1),wb.datemode)
       global_datetime=datetime.datetime(global_date[0],global_date[1],global_date[2],global_time[3],global_time[4],global_time[5])

    if ws.ncols/3>max:
     
     for c in range(max,ws.ncols/3):
      animal.append(list())
     
      ani=ws.cell_value(i,c*3).find('Turns')
      val=ws.cell_value(i,c*3)[:ani]
      id.append(val)
     
      from_time.append(list())
    
      flag.append(0)
      flag1.append(0)
      to_time.append(list())
    
      diff.append('')
      notes.append(list())
      notes2.append(list())
   
      notes1.append(list())

      start_date.append(0)
      start_time.append(0)
      start_datetime.append(0)
      end_date.append(0)
      end_time.append(0)
      end_datetime.append(0)
      sum.append(0)
      start_date[c]= global_date
      start_time[c]= global_time
     
      start_datetime[c]=global_datetime
      from_time[c].append(start_datetime[c])
     
     max=ws.ncols/3
   
    new_start_date=[0]*(ws.ncols/3)
    new_start_time=[0]*(ws.ncols/3)
    new_start_datetime=[0]*(ws.ncols/3)
   
       
       
        


    
    
   
    
     
    for row in range(i+1, ws.nrows):
      for col in range(0,ws.ncols/3):
         if end_datetime[col]!=0:
          old_date=end_datetime[col]
         else:
            old_date=0
         end_date[col]= xlrd.xldate_as_tuple(ws.cell_value(row,col*3),wb.datemode)
         end_time[col]= xlrd.xldate_as_tuple(ws.cell_value(row,col*3+1),wb.datemode)
         end_datetime[col]=datetime.datetime(end_date[col][0],end_date[col][1],end_date[col][2],end_time[col][3],end_time[col][4],end_time[col][5])
         if old_date==0 and start_datetime[col]!=end_datetime[col]:
           old_date=start_datetime[col]-datetime.timedelta(minutes=1)
           temp_oldate=old_date+datetime.timedelta(minutes=1)
           
           while end_datetime[col]>temp_oldate:
             
             
             diff[col]=('Gap from '+str(old_date+datetime.timedelta(minutes=1))+' to '+str(temp_oldate)+'. ')
             temp_oldate=temp_oldate+datetime.timedelta(minutes=1)
             if ((temp_oldate-start_datetime[col])>=datetime.timedelta(minutes=interval*t)):
               start_datetime[col]=start_datetime[col]+datetime.timedelta(minutes=interval*t)
               to_time[col].append(start_datetime[col]-datetime.timedelta(minutes=1))
               notes1[col].append(diff[col])
               if flag[col]==0:
                 notes[col].append('')
               else:
                    notes[col].append('NaN value(s) found. ')
               if flag1[col]==0:
                 notes2[col].append('')
        
               else:
                  notes2[col].append('Found '+str(flag1[col])+' value(s) greater than 126.')
               from_time[col].append(start_datetime[col])
               flag[col]=0
               flag1[col]=0
               animal[col].append(sum[col])
               sum[col]=0
               
               old_date=start_datetime[col]-datetime.timedelta(minutes=1)
               diff[col]=''
         if old_date!=0 and (old_date+datetime.timedelta(minutes=1))!=end_datetime[col]:
            temp_oldate=old_date+datetime.timedelta(minutes=1)
            while end_datetime[col]>temp_oldate:
             
             diff[col]=('Gap from '+str(old_date+datetime.timedelta(minutes=1))+' to '+str(temp_oldate)+'. ')
             temp_oldate=temp_oldate+datetime.timedelta(minutes=1)
             if ((temp_oldate-start_datetime[col])>=datetime.timedelta(minutes=interval*t)):
               start_datetime[col]=start_datetime[col]+datetime.timedelta(minutes=interval*t)
               to_time[col].append(start_datetime[col]-datetime.timedelta(minutes=1))
               notes1[col].append(diff[col])
               if flag[col]==0:
                 notes[col].append('')
               else:
                    notes[col].append('NaN value(s) found. ')
               if flag1[col]==0:
                 notes2[col].append('')
        
               else:
                  notes2[col].append('Found '+str(flag1[col])+' value(s) greater than 126.')
               from_time[col].append(start_datetime[col])
               flag[col]=0
               flag1[col]=0
               animal[col].append(sum[col])
               sum[col]=0
               
               old_date=start_datetime[col]-datetime.timedelta(minutes=1)
               diff[col]=''
         if ( (end_datetime[col]-start_datetime[col])>=datetime.timedelta(minutes=interval*t)):
         
           start_datetime[col]=start_datetime[col]+datetime.timedelta(minutes=interval*t)
           to_time[col].append(start_datetime[col]-datetime.timedelta(minutes=1))
         
           if diff[col]!='':
           
            
            notes1[col].append(diff[col])
            diff[col]=''
           else:
                    notes1[col].append('')
           from_time[col].append(start_datetime[col])
           
           
              
           animal[col].append(sum[col]) 
           if flag[col]==0:
            notes[col].append('') 
           else:
             notes[col].append('NaN value(s) found. ')
           if flag1[col]==0:
               notes2[col].append('')

           else:
                       notes2[col].append('Found '+str(flag1[col])+' value(s) greater than 126.')
           sum[col]=0
           flag[col]=0
           flag1[col]=0
      
         if ws.cell_value(row,col*3+2)!="NaN" :
                 
                  
                     
   			
                      
                  sum[col]=sum[col]+(int)(ws.cell_value(row,col*3+2))
         elif ws.cell_value(row,col*3+2)=="NaN" :
           flag[col]=1
         if ws.cell_value(row,col*3+2)!="NaN" and (int)(ws.cell_value(row,col*3+2))>=(int)(127):
           
           flag1[col]=flag1[col]+1

 for col in range(0,max):
     
     last_start.append(start_datetime[col])
     last_end.append(end_datetime[col])
 
 i=-1
 for col in range(0,max):
      i=i+1
      animal[i].append(sum[i])
      to_time[i].append(last_start[i]+datetime.timedelta(minutes=interval*t-1))
      if last_start[i]+datetime.timedelta(minutes=interval*t-1)!=last_end[i]:
       
        
        
        diff[col]+='Data missing for '+str(last_start[i]+datetime.timedelta(minutes=interval*t)-last_end[i]-datetime.timedelta(minutes=2))+'. '
        
        notes1[i].append(diff[col])
      else:
       
        notes1[i].append('')
      if flag[i]==0:
            notes[i].append('')
      else:
             notes[i].append('NaN value(s) found.')
      if flag1[i]==0:
            notes2[i].append('')
      else:
            notes2[i].append('Found '+str(flag1[i])+' value(s) greater than 126')



 for i in range(0,max):
     j=0
     for col1 in notes[i]:
     
      notes[i][j]=notes[i][j]+'\n'+notes2[i][j]+'\n'+notes1[i][j]
      j=j+1


     
 ofile=ofile.replace('"','')
 if os.path.isfile(ofile):
    wb=openpyxl.load_workbook(ofile)
 else:
    wb=openpyxl.Workbook()
 sheet = wb.get_sheet_names()


 wb.remove_sheet(wb.worksheets[0])
 ws=wb.create_sheet(0)
 ws.title='Wheel running'

 r=1
 Title=["Animal ID","From","To","Number of revolutions","Distance travelled (m)","Notes"]
 for t in range(1,len(Title)+1):
    
    ws.cell(row=r,column=t,value=Title[t-1])
 r=r+1
 i=0

 for c in range(0,max):

  ws.cell(row=r,column=1,value=id[i])
  j=0
  for t in to_time[i]:
   
   ws.cell(row=r,column=2,value=datetime.datetime.strftime(from_time[i][j],"%m/%d/%Y %H:%M:%S"))
   ws.cell(row=r,column=3,value=datetime.datetime.strftime(t,"%m/%d/%Y %H:%M:%S") )
   ws.cell(row=r,column=4,value=animal[i][j])
   ws.cell(row=r,column=5,value=animal[i][j]*0.2286*math.pi)

   ws.cell(row=r,column=6,value=notes[i][j])
   j=j+1
   r=r+1
  i=i+1
  r=r+1

 wb.save(ofile)





if __name__=='__main__':
      ofile=raw_input('Enter output filename \t')
      ofile=ofile.strip()
      interval=raw_input('Enter time interval for calculation \t')
      main(sys.argv[1:],ofile,interval )